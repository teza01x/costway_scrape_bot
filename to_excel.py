from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
from config import *


def create_new_excel():
    workbook = Workbook()
    sheet = workbook.active

    sheet.append(['ITEM NUMBER', 'PRICE', 'STOCK STATUS', 'TITLE', 'LINK'])

    now = datetime.now()
    current_excel_name = now.strftime("%d_%m_%y_%H.%M")
    workbook.save("out/{}.xlsx".format(current_excel_name))
    return "out/{}.xlsx".format(current_excel_name)


def add_new_data(data, excel_name):
    workbook = load_workbook(excel_name)

    sheet = workbook.active

    for i in data:
        sheet.append(i)

    try:
        workbook.save(excel_name)
    except Exception as e:
        print("Error {}".format(e))


def add_data_with_stock_and_price_change(data, excel_name, color_price, color_stock):
    workbook = load_workbook(excel_name)
    sheet = workbook.active

    fill1 = PatternFill(start_color=color_price, end_color=color_price, fill_type='solid')
    fill2 = PatternFill(start_color=color_stock, end_color=color_stock, fill_type='solid')

    for row_data in data:
        sheet.append(row_data)
        cell_to_color1 = sheet.cell(row=sheet.max_row, column=2)
        cell_to_color1.fill = fill1

        cell_to_color2 = sheet.cell(row=sheet.max_row, column=3)
        cell_to_color2.fill = fill2

    try:
        workbook.save(excel_name)
    except Exception as e:
        print("Error: {}".format(e))


def add_data_with_stock_change(data, excel_name, current_color):
    workbook = load_workbook(excel_name)
    sheet = workbook.active

    fill = PatternFill(start_color=current_color, end_color=current_color, fill_type='solid')

    for row_data in data:
        sheet.append(row_data)
        cell_to_color = sheet.cell(row=sheet.max_row, column=3)
        cell_to_color.fill = fill

    try:
        workbook.save(excel_name)
    except Exception as e:
        print("Error: {}".format(e))



def add_data_with_price_change(data, excel_name, current_color):
    workbook = load_workbook(excel_name)
    sheet = workbook.active

    fill = PatternFill(start_color=current_color, end_color=current_color, fill_type='solid')

    for row_data in data:
        sheet.append(row_data)
        cell_to_color = sheet.cell(row=sheet.max_row, column=2)
        cell_to_color.fill = fill

    try:
        workbook.save(excel_name)
    except Exception as e:
        print("Error: {}".format(e))
