import time
import warnings
import pandas as pd
import telebot
import sqlite3
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from config import *


class Objects:
    def __init__(self, wait):
        self.wait = wait


    def check_for_general_info(self):
        try:
            self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#wrapper > div.detail-wrapper > div.content > div.details-main > div.product-info-main')))
            return True
        except:
            return False


    def instock_status(self):
        try:
            self.wait.until(EC.element_to_be_clickable((By.XPATH, '//button[@class="ant-btn add"]')))
            return True
        except:
            return False


    def parse_info_from_page(self, browser):
        html = browser.page_source
        soup = BeautifulSoup(html, 'html.parser')
        right_menu = soup.find('div', class_='product-info-main')

        try:
            title = right_menu.find('h1', {'data-v-0f15b5a8': ''}).text.strip()
        except:
            title = 'None'
        try:
            price_element = right_menu.find('div', class_='price-row').find('span').text.strip()
            price = price_element.replace('C$', '')
            price = price.replace(',', '')
            price = float(price)
        except:
            price = 'None'
        try:
            item_number_element = right_menu.find('div', class_='item-no').text.strip()
            item_number_element = item_number_element.replace('Item No: ', '')
        except:
            item_number_element = 'None'
        return title, price, item_number_element


def create_new_excel():
    workbook = Workbook()
    sheet = workbook.active

    sheet.append(['ITEM NUMBER', 'PRICE', 'STOCK STATUS', 'TITLE', 'LINK'])

    now = datetime.now()
    current_excel_name = now.strftime("%d_%m_%y_%H.%M")
    workbook.save("out/{}.xlsx".format(current_excel_name))
    return "out/{}.xlsx".format(current_excel_name)


def add_new_data(data, excel_name):
    workbook = load_workbook(excel_name)

    sheet = workbook.active

    for i in data:
        sheet.append(i)

    try:
        workbook.save(excel_name)
    except Exception as e:
        print("Error {}".format(e))


def add_data_with_stock_and_price_change(data, excel_name, color_price, color_stock):
    workbook = load_workbook(excel_name)
    sheet = workbook.active

    fill1 = PatternFill(start_color=color_price, end_color=color_price, fill_type='solid')
    fill2 = PatternFill(start_color=color_stock, end_color=color_stock, fill_type='solid')

    for row_data in data:
        sheet.append(row_data)
        cell_to_color1 = sheet.cell(row=sheet.max_row, column=2)
        cell_to_color1.fill = fill1

        cell_to_color2 = sheet.cell(row=sheet.max_row, column=3)
        cell_to_color2.fill = fill2

    try:
        workbook.save(excel_name)
    except Exception as e:
        print("Error: {}".format(e))


def add_data_with_stock_change(data, excel_name, current_color):
    workbook = load_workbook(excel_name)
    sheet = workbook.active

    fill = PatternFill(start_color=current_color, end_color=current_color, fill_type='solid')

    for row_data in data:
        sheet.append(row_data)
        cell_to_color = sheet.cell(row=sheet.max_row, column=3)
        cell_to_color.fill = fill

    try:
        workbook.save(excel_name)
    except Exception as e:
        print("Error: {}".format(e))



def add_data_with_price_change(data, excel_name, current_color):
    workbook = load_workbook(excel_name)
    sheet = workbook.active

    fill = PatternFill(start_color=current_color, end_color=current_color, fill_type='solid')

    for row_data in data:
        sheet.append(row_data)
        cell_to_color = sheet.cell(row=sheet.max_row, column=2)
        cell_to_color.fill = fill

    try:
        workbook.save(excel_name)
    except Exception as e:
        print("Error: {}".format(e))


def item_exists(sku):
    conn = sqlite3.connect('items.db')
    cursor = conn.cursor()

    result = cursor.execute("SELECT sku FROM item_info WHERE sku = ?", (sku,))
    exists = bool(len(result.fetchall()))

    conn.close()

    return exists


def add_item_to_db(sku, price, stock):
    conn = sqlite3.connect('items.db')
    cursor = conn.cursor()

    cursor.execute("INSERT INTO item_info (sku, price, stock) VALUES(?, ?, ?)", (sku, price, stock,))

    conn.commit()
    conn.close()


def get_item_info(sku):
    conn = sqlite3.connect('items.db')
    cursor = conn.cursor()

    result = cursor.execute("SELECT price, stock FROM item_info WHERE sku = ?", (sku,))
    result = result.fetchall()[0]
    conn.close()

    return result


def update_item_info(sku, stock_status, item_price):
    conn = sqlite3.connect('items.db')
    cursor = conn.cursor()

    cursor.execute("UPDATE item_info SET price = ?, stock = ? WHERE sku = ?", (item_price, stock_status, sku,))

    conn.commit()
    conn.close()



def telegram_msg(bot_status):
    bot = telebot.TeleBot(telegram_token)
    if bot_status == "start":
        bot.send_message(telegram_user_id, "🟢 The CostWaybot started scraping pages. 🟢")
    elif bot_status == "end":
        bot.send_message(telegram_user_id, '♦️ The CostWaybot has finished parsing pages. ♦️\nThe bot has stopped working️, now you can work with excel spreadsheets. (folder "out")')


def main(browser, link, current_excel_name):
    wait = WebDriverWait(browser, 10)
    objects = Objects(wait)
    try:
        if objects.check_for_general_info() == True:
            item_info_list = list()
            item_title, item_price, item_sku = objects.parse_info_from_page(browser)
            if objects.instock_status() == True:
                item_stock = 'IN STOCK'
            else:
                item_stock = 'OUT OF STOCK'
            item_info_list.append(item_sku)
            item_info_list.append(item_price)
            item_info_list.append(item_stock)
            item_info_list.append(item_title)
            item_info_list.append(link)

            if item_exists(link) == False:
                add_item_to_db(link, item_price, item_stock)
                add_new_data([item_info_list], current_excel_name)
            elif item_exists(link) == True:
                prev_item_price, prev_item_stock = get_item_info(link)
                if prev_item_price == item_price and prev_item_stock == item_stock:
                    add_new_data([item_info_list], current_excel_name)
                elif prev_item_stock != item_stock and prev_item_price == item_price:
                    add_data_with_stock_change([item_info_list], current_excel_name, red_color)
                    update_item_info(link, item_stock, item_price)
                elif prev_item_stock == item_stock and prev_item_price != item_price:
                    add_data_with_price_change([item_info_list], current_excel_name, yellow_color)
                    update_item_info(link, item_stock, item_price)
                elif prev_item_stock != item_stock and prev_item_price != item_price:
                    add_data_with_stock_and_price_change([item_info_list], current_excel_name, yellow_color, red_color)
                    update_item_info(link, item_stock, item_price)
            time.sleep(5)
    except:
        time.sleep(10)


if __name__ == "__main__":
    service = Service()
    options = webdriver.ChromeOptions()
    warnings.filterwarnings("ignore", category=DeprecationWarning)

    try:
        browser = webdriver.Chrome(service=service, options=options)
        browser.maximize_window()

        current_excel_name = create_new_excel()

        df = pd.read_excel(aosom_link_excel_file)
        link_column = df['Supplier Product Link']

        telegram_msg("start")

        for i in range(len(link_column)):
            link = link_column.iloc[i]
            browser.get(link)
            main(browser, link, current_excel_name)

        browser.quit()
        telegram_msg("end")
    except:
        telegram_msg("end")